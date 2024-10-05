# Для работы со множеством типов файла
from hachoir.core.error import error
from hachoir.parser import createParser
from hachoir.core.tools import makePrintable
from hachoir.metadata import extractMetadata
# Для работы с изображениями
import exifread
from PIL import Image
import piexif
# Для работы с аудиофайлами
import audio_metadata
import eyed3
# Для работы с pdf-файлами
from PyPDF2 import PdfFileReader, PdfFileWriter
from PyPDF2.generic import NameObject, createStringObject
# Для работы с архивами и офисными документами
import datetime
from xml.etree import ElementTree as etree
import zipfile
# Для работы с офисными документами
import docx
import pptx
from openpyxl import load_workbook

class ImageMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:":
                list_meta.append(line[2:])
                i += 1
        return list_meta

    def extract_exif(self, path, metaType):
        list_meta = {}
        with open(path,"rb") as f:
            tags = exifread.process_file(f)
            for k,v in tags.items():
                if ('Thumbnail' in k and k != "JPEGThumbnail" and (metaType == 'All' or metaType == '0th' or metaType == '')):
                    k = k[10:]
                    list_meta['0th:' + k] = v
            for k, v in tags.items():
                if ('EXIF' in k  and (metaType == 'All' or metaType == 'Exif' or metaType == '')):
                    k = k[5:]
                    list_meta['Exif:' + k] = v
            for k, v in tags.items():
                if ('Image' in k and (metaType == 'All' or metaType == '1th' or metaType == '')):
                    k = k[6:]
                    list_meta['1th:' + k] = v
            f.close()
        print(list_meta)
        return list_meta

    def update_exif(self, path, attr, value):
        img = Image.open(path)
        exif_dict = piexif.load(img.info["exif"])
        result = 0
        try:
            # Изменение 0th метаданных
            if (attr == 'ImageWidth'):
                exif_dict['0th'][piexif.ImageIFD.ImageWidth] = int(value)
                result = 1
            elif (attr == 'ImageLength'):
                exif_dict['0th'][piexif.ImageIFD.ImageLength] = int(value)
                result = 1
            elif (attr == 'BitsPerSample'):
                value = value.split(',')
                value[0] = int(value[0])
                value[1] = int(value[1])
                value[2] = int(value[2])
                exif_dict['0th'][piexif.ImageIFD.BitsPerSample] = value
                result = 1
            elif (attr == 'Compression'):
                exif_dict['0th'][piexif.ImageIFD.Compression] = int(value)
                result = 1
            elif (attr == 'PhotometricInterpretation'):
                exif_dict['0th'][piexif.ImageIFD.PhotometricInterpretation] = int(value)
                result = 1
            elif (attr == 'Make'):
                exif_dict['0th'][piexif.ImageIFD.Make] = bytes(value)
                result = 1
            elif (attr == 'Model'):
                exif_dict['0th'][piexif.ImageIFD.Model] = bytes(value)
                result = 1
            elif (attr == 'Orientation'):
                exif_dict['0th'][piexif.ImageIFD.Orientation] = int(value)
                result = 1
            elif (attr == 'SamplesPerPixel'):
                exif_dict['0th'][piexif.ImageIFD.SamplesPerPixel] = int(value)
                result = 1
            elif (attr == 'XResolution'):
                exif_dict['0th'][piexif.ImageIFD.XResolution] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'YResolution'):
                exif_dict['0th'][piexif.ImageIFD.YResolution] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'PlanarConfiguration'):
                exif_dict['0th'][piexif.ImageIFD.PlanarConfiguration] = int(value)
                result = 1
            elif (attr == 'ResolutionUnit'):
                exif_dict['0th'][piexif.ImageIFD.ResolutionUnit] = int(value)
                result = 1
            elif (attr == 'Software'):
                exif_dict['0th'][piexif.ImageIFD.Software] = bytes(value)
                result = 1
            elif (attr == 'DateTime'):
                exif_dict['0th'][piexif.ImageIFD.DateTime] = bytes(value)
                result = 1
            elif (attr == 'Artist'):
                exif_dict['0th'][piexif.ImageIFD.Artist] = bytes(value)
                result = 1
            elif (attr == 'ExifOffset'):
                exif_dict['0th'][piexif.ImageIFD.ExifTag] = int(value)
                result = 1
            # Изменение Exif метаданных
            elif (attr == 'ExposureTime'):
                exif_dict['Exif'][piexif.ExifIFD.ExposureTime] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'FNumber'):
                exif_dict['Exif'][piexif.ExifIFD.FNumber] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'ISOSpeedRatings'):
                exif_dict['Exif'][piexif.ExifIFD.ISOSpeedRatings] = int(value)
                result = 1
            elif (attr == 'SensitivityType'):
                exif_dict['Exif'][piexif.ExifIFD.SensitivityType] = int(value)
                result = 1
            elif (attr == 'ExifVersion'):
                exif_dict['Exif'][piexif.ExifIFD.ExifVersion] = bytes(value)
                result = 1
            elif (attr == 'DateTimeOriginal'):
                exif_dict['Exif'][piexif.ExifIFD.DateTimeOriginal] = bytes(value)
                result = 1
            elif (attr == 'DateTimeDigitized'):
                exif_dict['Exif'][piexif.ExifIFD.DateTimeDigitized] = bytes(value)
                result = 1
            elif (attr == 'ShutterSpeedValue'):
                exif_dict['Exif'][piexif.ExifIFD.ShutterSpeedValue] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'ApertureValue'):
                exif_dict['Exif'][piexif.ExifIFD.ApertureValue] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'ExposureBiasValue'):
                exif_dict['Exif'][piexif.ExifIFD.ExposureBiasValue] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'MaxApertureValue'):
                exif_dict['Exif'][piexif.ExifIFD.MaxApertureValue] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'MeteringMode'):
                exif_dict['Exif'][piexif.ExifIFD.MeteringMode] = int(value)
                result = 1
            elif (attr == 'LightSource'):
                exif_dict['Exif'][piexif.ExifIFD.LightSource] = int(value)
                result = 1
            elif (attr == 'Flash'):
                exif_dict['Exif'][piexif.ExifIFD.Flash] = int(value)
                result = 1
            elif (attr == 'FocalLength'):
                exif_dict['Exif'][piexif.ExifIFD.FocalLength] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'SubSecTime'):
                exif_dict['Exif'][piexif.ExifIFD.SubSecTime] = bytes(value)
                result = 1
            elif (attr == 'SubSecTimeOriginal'):
                exif_dict['Exif'][piexif.ExifIFD.SubSecTimeOriginal] = bytes(value)
                result = 1
            elif (attr == 'SubSecTimeDigitized'):
                exif_dict['Exif'][piexif.ExifIFD.SubSecTimeDigitized] = bytes(value)
                result = 1
            elif (attr == 'ColorSpace'):
                exif_dict['Exif'][piexif.ExifIFD.ColorSpace] = int(value)
                result = 1
            elif (attr == 'ExifImageWidth'):
                exif_dict['Exif'][piexif.ExifIFD.PixelXDimension] = int(value)
                result = 1
            elif (attr == 'ExifImageLength'):
                exif_dict['Exif'][piexif.ExifIFD.PixelYDimension] = int(value)
                result = 1
            elif (attr == 'FocalPlaneXResolution'):
                exif_dict['Exif'][piexif.ExifIFD.FocalPlaneXResolution] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'FocalPlaneYResolution'):
                exif_dict['Exif'][piexif.ExifIFD.FocalPlaneYResolution] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'FocalPlaneResolutionUnit'):
                exif_dict['Exif'][piexif.ExifIFD.FocalPlaneResolutionUnit] = int(value)
                result = 1
            elif (attr == 'SensingMethod'):
                exif_dict['Exif'][piexif.ExifIFD.SensingMethod] = int(value)
                result = 1
            elif (attr == 'FileSource'):
                exif_dict['Exif'][piexif.ExifIFD.FileSource] = bytes(hex(value))
                result = 1
            elif (attr == 'SceneType'):
                exif_dict['Exif'][piexif.ExifIFD.SceneType] = bytes(hex(value))
                result = 1
            elif (attr == 'CFAPattern'):
                value = value.split(',')
                value[0] = int(value[0])
                value[1] = int(value[1])
                value[2] = int(value[2])
                value[3] = int(value[3])
                value[4] = int(value[4])
                value[5] = int(value[5])
                value[6] = int(value[6])
                value[7] = int(value[7])
                exif_dict['Exif'][piexif.ExifIFD.CFAPattern] = bytes(value)
                result = 1
            elif (attr == 'CustomRendered'):
                exif_dict['Exif'][piexif.ExifIFD.CustomRendered] = int(value)
                result = 1
            elif (attr == 'ExposureMode'):
                exif_dict['Exif'][piexif.ExifIFD.ExposureMode] = int(value)
                result = 1
            elif (attr == 'WhiteBalance'):
                exif_dict['Exif'][piexif.ExifIFD.WhiteBalance] = int(value)
                result = 1
            elif (attr == 'DigitalZoomRatio'):
                exif_dict['Exif'][piexif.ExifIFD.DigitalZoomRatio] = (int(value[:value.find('/')]), int(value[value.find('/')+1:]))
                result = 1
            elif (attr == 'FocalLengthIn35mmFilm'):
                exif_dict['Exif'][piexif.ExifIFD.FocalLengthIn35mmFilm] = int(value)
                result = 1
            elif (attr == 'SceneCaptureType'):
                exif_dict['Exif'][piexif.ExifIFD.SceneCaptureType] = int(value)
                result = 1
            elif (attr == 'GainControl'):
                exif_dict['Exif'][piexif.ExifIFD.GainControl] = int(value)
                result = 1
            elif (attr == 'Contrast'):
                exif_dict['Exif'][piexif.ExifIFD.Contrast] = int(value)
                result = 1
            elif (attr == 'Saturation'):
                exif_dict['Exif'][piexif.ExifIFD.Saturation] = int(value)
                result = 1
            elif (attr == 'Sharpness'):
                exif_dict['Exif'][piexif.ExifIFD.Sharpness] = int(value)
                result = 1
            elif (attr == 'SubjectDistanceRange'):
                exif_dict['Exif'][piexif.ExifIFD.SubjectDistanceRange] = int(value)
                result = 1
            elif (attr == 'BodySerialNumber'):
                exif_dict['Exif'][piexif.ExifIFD.BodySerialNumber] = bytes(value)
                result = 1
            elif (attr == 'LensSpecification'):
                value = value.split(',')
                value[0] = value[0].split('/')
                value[1] = value[1].split('/')
                value[2] = value[2].split('/')
                value[3] = value[3].split('/')
                exif_dict['Exif'][piexif.ExifIFD.LensSpecification] = ((int(value[0][0]),int(value[0][1])), (int(value[1][0]),int(value[1][1])), (int(value[2][0]),int(value[2][1])), (int(value[3][0]),int(value[3][1])))
                result = 1
            elif (attr == 'LensModel'):
                exif_dict['Exif'][piexif.ExifIFD.LensModel] = bytes(value)
                result = 1
            exif_bytes = piexif.dump(exif_dict)
            img.save(path, "jpeg", exif=exif_bytes)
        except:
            result = -1
        return result

    def remove_exif_all(self, path):
        piexif.remove(path)

class AudioMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:":
                list_meta.append(line[2:])
                i += 1
        return list_meta
    
    def extract_id3(self, path, metaType):
        list_meta = {}
        #audiofile = eyed3.load(path)
        #vers = audiofile.tag.version
        meta = audio_metadata.load(path)
        if (metaType == 'All' or metaType == 'MPEG' or metaType == ''):
            for k, v in meta.streaminfo.items():
                if (k.find('_') == -1):
                    list_meta['MPEG:' + k.title()] = v
        if (metaType == 'All' or 'ID3' in metaType or metaType == ''):
            for k, v in meta.tags.items():
                if (k.find('_') != 0):
                    if(k != "comment"):
                        #str(vers[0]) + "." + str(vers[1]) + "." + str(vers[2]) + 
                        list_meta['ID3:' + k.title()] = str(v)[2:-2]
                    else:
                        com = ""
                        for name in meta.tags.comment:
                            com = name
                        for k, v in com.items():
                            if (k.find('_') != 0):
                                # ' + str(vers[0]) + "." + str(vers[1]) + "." + str(vers[2]) + 
                                list_meta['ID3:' + k.title()] = v
        return list_meta
        
    def update_id3(self, path, attr, value):
        audiofile = eyed3.load(path)
        result = 0
        try:
            # Изменение MPEG метаданных
            if (attr == 'Layer'):
                audiofile.tag.track_num = int(value)
                result = 1
            # Изменение ID3 метаданных
            elif (attr == 'Title'):
                audiofile.tag.title = value
                result = 1
            elif (attr == 'Artist'):
                audiofile.tag.artist = value
                result = 1
            elif (attr == 'Album'):
                audiofile.tag.album = value
                result = 1
            elif (attr == 'Albumartist'):
                audiofile.tag.album_artist = value
                result = 1
            elif (attr == 'Genre'):
                audiofile.tag.genre = value
                result = 1
            audiofile.tag.save()
        except:
            result = -1
        return result

    def remove_id3_all(self, path):
        audiofile = eyed3.load(path)
        audiofile.tag.remove(path)

class VideoMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:" :
                if (line == 'Common:' or line == 'Video stream:' or line == 'Audio stream:'):
                    list_meta.append(line)
                else:
                    list_meta.append(line[2:])
                i += 1
        return list_meta

class PDFMeta():
    def extract_pdf(self, path):
        list_meta = {}
        pdf_file = PdfFileReader(path)
        pdf = pdf_file.getDocumentInfo()
        for data in pdf:
            list_meta[data[1:]] = pdf[data]
        return list_meta
    
    def extract_xmp(self, path, metaType):
        list_meta = {}
        pdf_file = PdfFileReader(path)
        xmpm = pdf_file.getXmpMetadata()
        xmp_methods = ['custom_properties', 'dc_contributor',
                        'dc_coverage', 'dc_creator',
                        'dc_date', 'dc_description',
                        'dc_format', 'dc_identifier',
                        'dc_language', 'dc_publisher',
                        'dc_relation', 'dc_rights',
                        'dc_source', 'dc_subject',
                        'dc_title', 'dc_type',
                        'pdf_keywords', 'pdf_pdfversion',
                        'pdf_producer', 'xmp_createDate',
                        'xmp_creatorTool', 'xmp_metadataDate',
                        'xmp_modifyDate', 'xmpmm_documentId',
                        'xmpmm_instanceId']
        xmpm_dict = {}
        for i in xmp_methods:
            try:
                if(metaType in i):
                    xmpm_dict[i] = getattr(xmpm, i)
            except:
                if(metaType in i):
                    xmpm_dict[i] = ''
        for data in xmpm_dict:
            if ('pdf_' in data):
                if (xmpm_dict[data] != ''):
                    list_meta['XMP-pdf:' + data[4:].title()] = xmpm_dict[data]
            elif ('dc_' in data):
                if isinstance(xmpm_dict[data], list):
                    if (len(xmpm_dict[data]) != 0):
                        list_meta['XMP-dc:' + data[4:].title()] = xmpm_dict[data][0]
                elif isinstance(xmpm_dict[data], dict):
                    if (len(xmpm_dict[data]) != 0):
                        list_meta['XMP-dc:' + data[4:].title()] = xmpm_dict[data]
            elif ('xmp_' in data):
                if (xmpm_dict[data] != ''):
                    list_meta['XMP-xmp:' + data[4:].title()] = xmpm_dict[data]
            elif ('xmpmm_' in data):
                if (xmpm_dict[data] != ''):
                    list_meta['XMP-xmpmm:' + data[6:].title()] = xmpm_dict[data]
        return list_meta

    def update_pdf(self, path, attr, value):
        result = 0
        list_meta = PDFMeta().extract_pdf(path)
        list_pdf = {}
        try:
            for k, v in list_meta.items():
                if (k in attr):
                    list_pdf['/' + k] = value
                    result = 1
                else:
                    list_pdf["/" + k] = v
            file_in = open(path, 'rb')
            pdf_reader = PdfFileReader(file_in)
            pdf_writer = PdfFileWriter()
            pdf_writer.appendPagesFromReader(pdf_reader)
            pdf_writer.addMetadata(list_pdf)
            file_in.close()
            file_out = open(path, 'wb')
            pdf_writer.write(file_out)
            file_out.close()
        except:
            result = -1
        return result

    def remove_pdf(self, path, attr):
        result = 0
        list_meta = PDFMeta().extract_pdf(path)
        list_pdf = {}
        for k, v in list_meta.items():
            if (k in attr):
                result = 1
            else:
                list_pdf["/" + k] = v
        file_in = open(path, 'rb')
        pdf_reader = PdfFileReader(file_in)
        pdf_writer = PdfFileWriter()
        pdf_writer.appendPagesFromReader(pdf_reader)
        pdf_writer.addMetadata(list_pdf)
        file_in.close()
        file_out = open(path, 'wb')
        pdf_writer.write(file_out)
        file_out.close()
        return result
    
    def remove_pdf_all(self, path):
        list_pdf = {}
        file_in = open(path, 'rb')
        pdf_reader = PdfFileReader(file_in)
        pdf_writer = PdfFileWriter()
        pdf_writer.appendPagesFromReader(pdf_reader)
        pdf_writer.addMetadata(list_pdf)
        file_in.close()
        file_out = open(path, 'wb')
        pdf_writer.write(file_out)
        file_out.close()
    
class TextMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:":
                list_meta.append(line[2:])
                i += 1
        return list_meta
        
class ArchivesMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:":
                if (line != 'Common:'):
                    list_meta.append(line)
        return list_meta

    def extract_zip(self, path):
        list_meta = []
        if (zipfile.is_zipfile(path)):
            zfile = zipfile.ZipFile(path)
            for key in zfile.infolist():
                list_meta.append("ZipFileName: " + str(key.filename))
                list_meta.append("ZipOrigName: " + str(key.orig_filename))
                list_meta.append("ZipRequiredVersion: " + str(key.extract_version))
                list_meta.append("ZipBitFlag: " + str(key.flag_bits))
                list_meta.append("ZipCompression: " + str(key.compress_type))
                list_meta.append("ZipFileMode: " + str(key.reserved))
                list_meta.append("ZipModifyDate: " + str(datetime.datetime(*key.date_time)))
                list_meta.append("ZipCreateSystem: " + str(key.create_system))
                list_meta.append("ZipCRC: " + str(key.CRC))
                list_meta.append("ZipCompressedSize: " + str(key.compress_size))
                list_meta.append("ZipComments: " + str(key.comment))
                list_meta.append("ZIP version: " + str(key.create_version))
        return list_meta

class OfficeMeta():
    def extract_zip(self, path):
        list_meta = []
        if (zipfile.is_zipfile(path)):
            zfile = zipfile.ZipFile(path)
            for key in zfile.infolist():
                list_meta.append("ZipFileName: " + str(key.filename))
                list_meta.append("ZipOrigName: " + str(key.orig_filename))
                list_meta.append("ZipRequiredVersion: " + str(key.extract_version))
                list_meta.append("ZipBitFlag: " + str(key.flag_bits))
                list_meta.append("ZipCompression: " + str(key.compress_type))
                list_meta.append("ZipFileMode: " + str(key.reserved))
                list_meta.append("ZipModifyDate: " + str(datetime.datetime(*key.date_time)))
                list_meta.append("ZipCreateSystem: " + str(key.create_system))
                list_meta.append("ZipCRC: " + str(key.CRC))
                list_meta.append("ZipCompressedSize: " + str(key.compress_size))
                list_meta.append("ZipComments: " + str(key.comment))
                list_meta.append("ZIP version: " + str(key.create_version))
        return list_meta

    def extract_docx(self, path):
        doc = docx.Document(path)
        metadata = {}
        prop = doc.core_properties
        metadata["XML-dc:Title"] = prop.title
        metadata["XML-dc:Subject"] = prop.subject
        metadata["XML-dc:Author"] = prop.author
        metadata["XML-dc:Comments"] = prop.comments
        metadata["XML:Category"] = prop.category
        metadata["XML:Content_Status"] = prop.content_status
        metadata["XML:Created"] = prop.created
        metadata["XML:Identifier"] = prop.identifier
        metadata["XML:Keywords"] = prop.keywords
        metadata["XML:Language"] = prop.language
        metadata["XML:Last_Modified"] = prop.last_modified_by
        metadata["XML:Last_Printed"] = prop.last_printed
        metadata["XML:Modified"] = prop.modified
        metadata["XML:Revision"] = prop.revision
        metadata["XML:Version"] = prop.version
        return metadata

    def extract_pptx(self, path):
        ppt = pptx.Presentation(path)
        metadata = {}
        prop = ppt.core_properties
        metadata["XML-dc:Title"] = prop.title
        metadata["XML-dc:Subject"] = prop.subject
        metadata["XML-dc:Author"] = prop.author
        metadata["XML-dc:Comments"] = prop.comments
        metadata["XML:Category"] = prop.category
        metadata["XML:Content_Status"] = prop.content_status
        metadata["XML:Created"] = prop.created
        metadata["XML:Identifier"] = prop.identifier
        metadata["XML:Keywords"] = prop.keywords
        metadata["XML:Language"] = prop.language
        metadata["XML:Last_Modified"] = prop.last_modified_by
        metadata["XML:Last_Printed"] = prop.last_printed
        metadata["XML:Modified"] = prop.modified
        metadata["XML:Revision"] = prop.revision
        metadata["XML:Version"] = prop.version
        return metadata
    
    def extract_xlsx(self, path):
        wb = load_workbook(path)
        metadata = {}
        prop = wb.properties
        metadata["XML-dc:Title"] = prop.title
        metadata["XML-dc:Subject"] = prop.subject
        metadata["XML-dc:Author"] = prop.author
        metadata["XML-dc:Comments"] = prop.comments
        metadata["XML:Category"] = prop.category
        metadata["XML:Content_Status"] = prop.content_status
        metadata["XML:Created"] = prop.created
        metadata["XML:Identifier"] = prop.identifier
        metadata["XML:Keywords"] = prop.keywords
        metadata["XML:Language"] = prop.language
        metadata["XML:Last_Modified"] = prop.last_modified_by
        metadata["XML:Last_Printed"] = prop.last_printed
        metadata["XML:Modified"] = prop.modified
        metadata["XML:Revision"] = prop.revision
        metadata["XML:Version"] = prop.version
        return metadata

    def extract_xml(self, path, ext, metaType):
        metadata = {}
        if (ext == ".doc" or ext == ".doc"):
            metadata = OfficeMeta().extract_docx(path)
        elif (ext == ".ppt" or ext == ".pptx"):
            metadata = OfficeMeta().extract_pptx(path)
        elif (ext == ".xls" or ext == ".xlsx"):
            metadata = OfficeMeta().extract_xlsx(path)
        isApp = False
        if (zipfile.is_zipfile(path)):
            zfile = zipfile.ZipFile(path)
            for i in zfile.namelist():
                if (i == 'docProps/app.xml'):
                    isApp = True
            if (isApp):
                app_xml = etree.fromstring(zfile.read('docProps/app.xml'))
                app_mapping = {
                'TotalTime': 'Edit Time (minutes)',
                'Pages': 'Page Count',
                'Words': 'Word Count',
                'Characters': 'Character Count',
                'Lines': 'Line Count',
                'Paragraphs': 'Paragraph Count',
                'Company': 'Company',
                'HyperlinkBase': 'Hyperlink Base',
                'Slides': 'Slide count',
                'Notes': 'Note Count',
                'HiddenSlides': 'Hidden Slide Count',
                }
                for element in app_xml.getchildren():
                    for key, title in app_mapping.items():
                        if key in element.tag:
                            if 'date' in title.lower():
                                text = datetime.datetime.strptime(element.text, "%Y-%m-%dT%H:%M:%SZ")
                            else:
                                text = element.text
                            metadata["XML:" + title.title()] = text
        zfile.close()
        if (metaType == "All"):
            return metadata
        else:
            if (metaType == "XML-dc"):
                list_dc = {}
                for k, v in metadata.items():
                    if ("XML-dc" in k):
                        list_dc[k] = v
                return list_dc
            else:
                list_xml = {}
                for k, v in metadata.items():
                    if ("XML" in k):
                        list_xml[k] = v
                return list_xml
    
    def update_docx(self, path, attr, value):
        doc = docx.Document(path)
        prop = doc.core_properties
        result = 0
        try:
            if (attr == 'Title'):
                prop.title = value
                result = 1
            elif (attr == 'Subject'):
                prop.subject = value
                result = 1
            elif (attr == 'Author'):
                prop.author = value
                result = 1
            elif (attr == 'Comments'):
                prop.comments = value
                result = 1
            elif (attr == 'Category'):
                prop.category = value
                result = 1
            elif (attr == 'Content_Status'):
                prop.content_status = value
                result = 1
            elif (attr == 'Created'):
                prop.created = datetime.datetime(*value)
                result = 1
            elif (attr == 'Keywords'):
                prop.keywords = value
                result = 1
            elif (attr == 'Language'):
                prop.language = value
                result = 1
            elif (attr == 'Last_Modified'):
                prop.last_modified_by = value
                result = 1
            elif (attr == 'Last_Printed'):
                prop.last_printed = value
                result = 1
            elif (attr == 'Modified'):
                prop.modified = datetime.datetime(*value)
                result = 1
            elif (attr == 'Revision'):
                prop.revision = int(value)
                result = 1
            elif (attr == 'Version'):
                prop.version = value
                result = 1
            doc.save(path)
        except:
            result = -1
        return result
    
    def update_pptx(self, path, attr, value):
        ppt = pptx.Presentation(path)
        prop = ppt.core_properties
        result = 0
        try:
            if (attr == 'Title'):
                prop.title = value
                result = 1
            elif (attr == 'Subject'):
                prop.subject = value
                result = 1
            elif (attr == 'Author'):
                prop.author = value
                result = 1
            elif (attr == 'Comments'):
                prop.comments = value
                result = 1
            elif (attr == 'Category'):
                prop.category = value
                result = 1
            elif (attr == 'Content_Status'):
                prop.content_status = value
                result = 1
            elif (attr == 'Created'):
                prop.created = datetime.datetime(*value)
                result = 1
            elif (attr == 'Keywords'):
                prop.keywords = value
                result = 1
            elif (attr == 'Language'):
                prop.language = value
                result = 1
            elif (attr == 'Last_Modified'):
                prop.last_modified_by = value
                result = 1
            elif (attr == 'Last_Printed'):
                prop.last_printed = value
                result = 1
            elif (attr == 'Modified'):
                prop.modified = datetime.datetime(*value)
                result = 1
            elif (attr == 'Revision'):
                prop.revision = int(value)
                result = 1
            elif (attr == 'Version'):
                prop.version = value
                result = 1
            ppt.save(path)
        except:
            result = -1
        return result

    def update_xlsx(self, path, attr, value):
        wb = load_workbook(path)
        prop = wb.properties
        result = 0
        try:
            if (attr == 'Title'):
                prop.title = value
                result = 1
            elif (attr == 'Subject'):
                prop.subject = value
                result = 1
            elif (attr == 'Author'):
                prop.author = value
                result = 1
            elif (attr == 'Comments'):
                prop.comments = value
                result = 1
            elif (attr == 'Category'):
                prop.category = value
                result = 1
            elif (attr == 'Content_Status'):
                prop.content_status = value
                result = 1
            elif (attr == 'Created'):
                prop.created = datetime.datetime(*value)
                result = 1
            elif (attr == 'Keywords'):
                prop.keywords = value
                result = 1
            elif (attr == 'Language'):
                prop.language = value
                result = 1
            elif (attr == 'Last_Modified'):
                prop.last_modified_by = value
                result = 1
            elif (attr == 'Last_Printed'):
                prop.last_printed = value
                result = 1
            elif (attr == 'Modified'):
                prop.modified = datetime.datetime(*value)
                result = 1
            elif (attr == 'Revision'):
                prop.revision = int(value)
                result = 1
            elif (attr == 'Version'):
                prop.version = value
                result = 1
            wb.save(path)
        except:
            result = -1
        return result

class EXEMeta():
    def extract_base(self, path):
        list_meta = []
        parser = createParser(path)
        if not parser:
            return list_meta
        try:
            metadata = extractMetadata(parser)
        except error as err:
            return list_meta
        if not metadata:
            return list_meta
        text = metadata.exportPlaintext()
        i = 0
        for line in text:
            if line != "Metadata:":
                list_meta.append(line[2:])
                i += 1
        return list_meta